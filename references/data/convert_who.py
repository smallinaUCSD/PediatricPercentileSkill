"""Regenerates who_*.csv from the raw WHO xlsx files in who_raw/.

Requires openpyxl (not a runtime dependency of the engine): `uv run
--with openpyxl python3 references/data/convert_who.py`.
"""

import csv
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
RAW = HERE / "who_raw"
OUT = HERE

# (indicator, axis_col_name_out, girls_file, boys_file, out_file)
JOBS = [
    ("weight_for_age", "age_days", "who_wfa_girls.xlsx", "who_wfa_boys.xlsx", "who_weight_for_age.csv"),
    ("length_for_age", "age_days", "who_lhfa_girls.xlsx", "who_lhfa_boys.xlsx", "who_length_for_age.csv"),
    ("head_circumference_for_age", "age_days", "who_hcfa_girls.xlsx", "who_hcfa_boys.xlsx", "who_head_circumference_for_age.csv"),
    ("bmi_for_age", "age_days", "who_bfa_girls.xlsx", "who_bfa_boys.xlsx", "who_bmi_for_age.csv"),
    ("weight_for_length", "length_cm", "who_wfl_girls.xlsx", "who_wfl_boys.xlsx", "who_weight_for_length.csv"),
]

for indicator, axis_name, girls_fn, boys_fn, out_fn in JOBS:
    rows = []
    for sex_code, fn in [(2, girls_fn), (1, boys_fn)]:
        wb = openpyxl.load_workbook(RAW / fn, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = row
                assert header[0] in ("Day", "Length", "Height"), header
                assert header[1:4] == ("L", "M", "S"), header
                continue
            axis_val, L, M, S = row[0], row[1], row[2], row[3]
            if axis_val is None:
                continue
            rows.append((sex_code, axis_val, L, M, S))
    with open(OUT / out_fn, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Sex", axis_name, "L", "M", "S"])
        for r in rows:
            w.writerow(r)
    print(out_fn, len(rows), "rows")
